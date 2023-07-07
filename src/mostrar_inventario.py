from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import mysql.connector
import tempfile
import base64
from flask_wtf.csrf import CSRFProtect

from werkzeug.utils import secure_filename
import os
import pandas as pd

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.drawing.image import Image

mostrar_inventario = Flask(__name__)
mostrar_inventario.config['UPLOAD_FOLDER'] = 'static/imgRepuestos'
mostrar_inventario.config['TEMPLATES_AUTO_RELOAD'] = True

db_config = {
    'user': 'root',
    'password': '',
    'host': 'localhost',
    'database': 'curpiscobd_repuesto'
}

mostrar_inventario.secret_key = 'mysecretkey'
csrf = CSRFProtect(mostrar_inventario)

@mostrar_inventario.route('/mostrar_inventario')
def index_mostrar_inventario():
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM registrar_repuesto')
    data = cur.fetchall()
    
    return render_template('Inventario.html', repuestos=data)

@mostrar_inventario.route('/limpiar_filtro', methods=['GET'])
def index_limpiar_filtro():
    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
    cur = conn.cursor()

    cur.execute('SELECT * FROM registrar_repuesto')
    data = cur.fetchall()

    conn.close()

    converted_data = []
    for value in data:
        converted_row = []
        for val in value:
            if isinstance(val, bytes):
                converted_row.append(val.decode('utf-8'))
            else:
                converted_row.append(val)
        converted_data.append(converted_row)

    return jsonify({'repuesto': converted_data})

@mostrar_inventario.route('/buscar_repuesto', methods=['POST'])
def index_buscar_repuesto():
    nombre_repuesto = request.form.get('nombre_repuesto')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
    cur = conn.cursor()

    cur.execute('SELECT * FROM registrar_repuesto WHERE nombre_repuesto LIKE %s', (f'%{nombre_repuesto}%',))
    data = cur.fetchall()

    conn.close()

    if not data:
        return jsonify({'repuesto': []})

    converted_data = []
    for row in data:
        converted_row = []
        for value in row:
            if isinstance(value, bytes):
                converted_row.append(value.decode('utf-8'))
            else:
                converted_row.append(value)
        converted_data.append(converted_row)

    return jsonify({'repuesto': converted_data})

@mostrar_inventario.route('/edit/<id>', methods=['POST'])
@csrf.exempt
def index_edit_inventario(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('SELECT id_repuesto, codigo_repuesto, nombre_repuesto, descripcion_repuesto, cantidad_minima_repuesto, foto_repuesto, cantidad_total FROM registrar_repuesto WHERE id_repuesto = %s', (id,))
    data = cur.fetchone()
    
    if data:
        id_repuesto, codigo_repuesto, nombre_repuesto, descripcion_repuesto, cantidad_minima_repuesto, foto_repuesto, cantidad_total = data

        if foto_repuesto is not None:
            foto_repuesto_encoded = base64.b64encode(foto_repuesto).decode('ascii')
        else:
            foto_repuesto_encoded = None

        response_data = {
            'id_repuesto':id_repuesto,
            'codigo_repuesto': codigo_repuesto,
            'nombre_repuesto': nombre_repuesto,
            'descripcion_repuesto': descripcion_repuesto,
            'cantidad_minima_repuesto': cantidad_minima_repuesto,
            'foto_repuesto': foto_repuesto_encoded,
            'cantidad_total': cantidad_total
        }
    else:
        response_data = {}
        
    cur.close()
    conn.close()

    print(response_data)
    
    return render_template('editInventario.html', repuesto=response_data)

@mostrar_inventario.route('/update/<id>', methods=['POST'])
def index_update_repuesto(id):
    if request.method == 'POST':
        codigo_repuesto = request.form['codigo_repuesto']
        nombre_repuesto = request.form['nombre_repuesto']
        descripcion_repuesto = request.form['descripcion_repuesto']
        cantidad_minima_repuesto = request.form['cantidad_minima_repuesto']
        foto_repuesto = request.files['foto_repuesto']
        cantidad_total = request.form['cantidad_total']

        conn = mysql.connector.connect(**db_config)
        cur = conn.cursor()
        cur.execute(
            'SELECT foto_repuesto FROM registrar_repuesto WHERE id_repuesto = %s', (id,))
        foto_existente = cur.fetchone()
        foto_repuesto_path = foto_existente[0] if foto_existente else ''

        nuevo_nombre_foto = ''  

        if foto_repuesto.filename:
            basepath = os.path.dirname(__file__)
            filename = secure_filename(foto_repuesto.filename)
            extension = os.path.splitext(filename)[1]
            nuevo_nombre_foto = nombre_repuesto + extension 
            upload_path = os.path.join(basepath, 'static/imgRepuestos', nuevo_nombre_foto)
            foto_repuesto.save(upload_path)
            foto_repuesto_path = upload_path

        cur.execute("""
            UPDATE registrar_repuesto
            SET codigo_repuesto = %s,
                nombre_repuesto = %s,
                descripcion_repuesto = %s,
                cantidad_minima_repuesto = %s,
                foto_repuesto = %s,
                cantidad_total = %s
            WHERE id_repuesto = %s
        """, (codigo_repuesto, nombre_repuesto, descripcion_repuesto, cantidad_minima_repuesto, nuevo_nombre_foto, cantidad_total, id))
        conn.commit()

        cur.close()
        conn.close()
        return redirect(url_for('mostrar_inventario'))

@mostrar_inventario.route('/delete_inventario/<string:id>', methods=['POST', 'GET'])
def index_delete_repuesto(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('DELETE FROM registrar_repuesto WHERE id_repuesto = %s', (id,))
    conn.commit()

    cur.close()
    conn.close()
    return redirect(url_for('mostrar_inventario'))

def obtener_datos_mysql():
    conn = mysql.connector.connect(**db_config)

    consulta = "SELECT codigo_repuesto, nombre_repuesto, descripcion_repuesto, cantidad_minima_repuesto, cantidad_total FROM registrar_repuesto"
    datos = pd.read_sql_query(consulta, conn)

    titulos_encabezado = ["CÓDIGO", "NOMBRE", "DESCRIPCIÓN", "CANTIDAD MINIMA", "CANTIDAD TOTAL"]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Datos'

    for c, titulo in enumerate(titulos_encabezado, start=1):
        cell = worksheet.cell(row=8, column=c)
        cell.value = titulo
        cell.font = Font(bold=True, color="FFFFFFFF", name="Times New Roman")
        cell.fill = PatternFill(fill_type='solid', fgColor='9a5833')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    for r, row in enumerate(datos.values, start=9):
        for c, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r, column=c)
            cell.value = value
            cell.font = Font(name='Times New Roman')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    for column_cells in worksheet.columns:
        column = column_cells[0].column_letter

        if column == 'A':
            worksheet.column_dimensions[column].width = 18
        elif column == 'B':
            worksheet.column_dimensions[column].width = 42
        elif column == 'C':
            worksheet.column_dimensions[column].width = 54
        elif column == 'D':
            worksheet.column_dimensions[column].width = 24
        elif column == 'E':
            worksheet.column_dimensions[column].width = 24

    for r, row_cells in enumerate(worksheet.iter_rows(), start=1):
        adjusted_height = 18
        for cell in row_cells:
            worksheet.row_dimensions[cell.row].height = adjusted_height

    imagen_path = 'src/static/img/ExcelEncabezadoRepuestos.png'
    imagen_width = 1136.12
    imagen_height = 166.67

    imagen = Image(imagen_path)
    imagen.width = imagen_width
    imagen.height = imagen_height

    worksheet.add_image(imagen, 'A1')

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        archivo_temporal = tmp.name
        workbook.save(archivo_temporal)

    conn.close()

    return archivo_temporal

@mostrar_inventario.route('/generar_excel', methods=['POST'])
def index_generar_excel():
    archivo_excel = obtener_datos_mysql()

    return send_file(archivo_excel, as_attachment=True)
    
if __name__ == '__main__':
    mostrar_inventario.run(port=3000, debug=True)
