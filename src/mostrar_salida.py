from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import mysql.connector
import tempfile
import base64

import pandas as pd

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.drawing.image import Image

mostrar_salida = Flask(__name__)

db_config = {
    'user': 'root',
    'password': '',
    'host': 'localhost',
    'database': 'curpiscobd_repuesto'
}

mostrar_salida.secret_key = 'mysecretkey'

@mostrar_salida.route('/')
def index_mostrar_salida():
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute("SELECT s.id_salida_repuesto, r.codigo_repuesto, r.nombre_repuesto, s.cantidad_salida_repuesto, t.apellidos_responsable, s.fecha_salida_repuesto "
            "FROM salida_repuesto s "
            "JOIN registrar_repuesto r ON s.id_repuesto = r.id_repuesto "
            "JOIN registrar_responsable t ON s.id_responsable = t.id_responsable")

    results = cur.fetchall()

    cur.close()
    conn.close()
    
    return render_template('MostrarSalida.html', repuestos=results)

@mostrar_salida.route('/buscar_salida', methods=['POST'])
def index_buscar_salida():
    codigo_repuesto = request.form.get('codigo_repuesto')
    fecha_salida = request.form.get('fecha_salida')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify ({'repuesto_salida': []})

    cur = conn.cursor()

    query = """
    SELECT s.id_salida_repuesto, r.codigo_repuesto, r.nombre_repuesto, s.cantidad_salida_repuesto, t.apellidos_responsable, DATE_FORMAT(s.fecha_salida_repuesto, '%Y-%m-%d')
    FROM salida_repuesto s
    JOIN registrar_repuesto r ON s.id_repuesto = r.id_repuesto
    JOIN registrar_responsable t ON s.id_responsable = t.id_responsable
    WHERE 1=1
    """
    params = []

    if codigo_repuesto:
        cur.execute('SELECT id_repuesto FROM registrar_repuesto WHERE codigo_repuesto = %s', (codigo_repuesto,))
        id_repuesto = cur.fetchone()
        if id_repuesto is None:
            conn.close()
            return jsonify({'repuesto_salida': []})

        id_repuesto = id_repuesto[0]
        query += "AND r.id_repuesto = %s "
        params.append(id_repuesto)

    if fecha_salida:
        query += "AND DATE(s.fecha_salida_repuesto) = %s "
        params.append(fecha_salida)

    cur.execute(query, params)
    data = cur.fetchall()

    conn.close()

    converted_data = []
    for row in data:
        converted_row = []
        for value in row:
            if isinstance(value, bytes):
                converted_row.append(value.decode('utf-8'))
            else:
                converted_row.append(value)
        converted_data.append(converted_row)

    return jsonify({'repuesto_salida': converted_data})

@mostrar_salida.route('/limpiar_filtro', methods=['GET'])
def index_limpiar_filtroS():
    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify({'repuesto': []})

    cur = conn.cursor()

    query = """
    SELECT s.id_salida_repuesto, r.codigo_repuesto, r.nombre_repuesto, s.cantidad_salida_repuesto, t.apellidos_responsable, DATE_FORMAT(s.fecha_salida_repuesto, '%Y-%m-%d')
    FROM salida_repuesto s
    JOIN registrar_repuesto r ON s.id_repuesto = r.id_repuesto
    JOIN registrar_responsable t ON s.id_responsable = t.id_responsable
    """

    cur.execute(query)
    data = cur.fetchall()

    conn.close()

    converted_data = []
    for row in data:
        converted_row = []
        for value in row:
            if isinstance(value, bytes):
                converted_row.append(value.decode('utf-8'))
            else:
                converted_row.append(value)
        converted_data.append(converted_row)

    return jsonify({'repuesto': converted_data})

@mostrar_salida.route('/edit_salida/<id>')
def index_edit_salida(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute("""
        SELECT s.id_salida_repuesto, r.codigo_repuesto, r.nombre_repuesto, r.descripcion_repuesto, r.foto_repuesto, s.cantidad_salida_repuesto, t.codigo_responsable, t.apellidos_responsable, DATE_FORMAT(s.fecha_salida_repuesto, '%Y-%m-%d')
        FROM salida_repuesto s
        JOIN registrar_repuesto r ON s.id_repuesto = r.id_repuesto
        JOIN registrar_responsable t ON s.id_responsable = t.id_responsable
        WHERE s.id_salida_repuesto = %s
    """, (id,))
    data = cur.fetchall()
    
    column_titles = ['id_repuesto', 'codigo_repuesto', 'nombre_repuesto', 'descripcion_repuesto', 'foto_repuesto', 'cantidad_salida_repuesto', 'codigo_responsable', 'apellidos_responsable', 'fecha_salida_repuesto']
    response_data = {title: [row[i] for row in data] for i, title in enumerate(column_titles)}
    response_data = {key: value[0] if isinstance(value, list) and len(value) == 1 else value for key, value in response_data.items()}
        
    
    return render_template('editSalida.html', repuesto=response_data)

@mostrar_salida.route('/autocompletar_salidaS', methods=['GET', 'POST'])
def autocompletar_repuesto_salidaS():
    codigo_repuesto = request.form.get('codigo_repuesto')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify({'data': {}})
    cursor = conn.cursor()

    query = "SELECT nombre_repuesto, descripcion_repuesto, foto_repuesto FROM registrar_repuesto WHERE codigo_repuesto = %s"
    cursor.execute(query, (codigo_repuesto,))
    result = cursor.fetchone()

    if result:
        nombre_repuesto, descripcion_repuesto, foto_repuesto = result

        if foto_repuesto is not None:
            foto_repuesto_encoded = base64.b64encode(foto_repuesto).decode('ascii')
        else:
            foto_repuesto_encoded = None

        response_data = {
            'nombre_repuesto': nombre_repuesto,
            'descripcion_repuesto': descripcion_repuesto,
            'foto_repuesto': foto_repuesto_encoded
        }
    else:
        response_data = {}

    cursor.close()
    conn.close()

    return jsonify({'data': response_data})

@mostrar_salida.route('/autocompletar_responsableS', methods=['GET', 'POST'])
def autocompletar_responsable_salidaS():
    codigo_responsable = request.form.get('codigo_responsable')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify({'data': {}}) 

    cursor = conn.cursor()
    query = "SELECT nombres_responsable, apellidos_responsable FROM registrar_responsable WHERE codigo_responsable = %s"
    cursor.execute(query, (codigo_responsable,))
    result = cursor.fetchone()
    print("Result es:", result)

    cursor.close()
    conn.close()

    data = {
        'nombres_responsable': result[0],
        'apellidos_responsable': result[1]
    }

    return jsonify({'data': data})

@mostrar_salida.route('/update_salida/<id>', methods=['POST'])
def index_update_salida(id):
    if request.method == 'POST':
        codigo_repuesto = request.form['codigo_repuesto']
        cantidad_salida_repuesto = int(request.form['cantidad_salida_repuesto'])
        codigo_responsable = request.form['codigo_responsable']
        fecha_salida_repuesto = request.form['fecha_salida_repuesto']
        
        conn = mysql.connector.connect(**db_config)
        if conn.is_connected():
            print("Conexión exitosa a la base de datos")
        else:
            print("No se pudo establecer la conexión a la base de datos")
        
        cur = conn.cursor()
        cur.execute('SELECT id_repuesto FROM registrar_repuesto WHERE codigo_repuesto = %s', (codigo_repuesto,))
        id_repuesto = cur.fetchone()    
        id_repuesto = id_repuesto[0] 
        cur.execute('SELECT id_responsable FROM registrar_responsable WHERE codigo_responsable = %s', (codigo_responsable,))
        id_responsable = cur.fetchone()    
        id_responsable = id_responsable[0] 
        
        cur.execute('SELECT cantidad_salida_repuesto FROM salida_repuesto WHERE id_salida_repuesto = %s', (id,))
        cantidad_antes = cur.fetchone()[0] 

        cur.execute('SELECT cantidad_total FROM registrar_repuesto WHERE id_repuesto = %s', (id_repuesto,))
        result = cur.fetchone()
        
        if result:
            cantidad_actual = result[0]
            aumentar = cantidad_actual + int(cantidad_antes)
            nueva_cantidad = aumentar - cantidad_salida_repuesto
            cur.execute('UPDATE registrar_repuesto SET cantidad_total = %s WHERE id_repuesto = %s',
            (nueva_cantidad, id_repuesto))
        
            cur.execute("""
                UPDATE salida_repuesto
                SET id_repuesto = %s,
                    cantidad_salida_repuesto = %s,
                    id_responsable = %s,
                    fecha_salida_repuesto = %s
                WHERE id_salida_repuesto  = %s
            """, (id_repuesto, cantidad_salida_repuesto, id_responsable, fecha_salida_repuesto, id))
            conn.commit()

        cur.close()
        conn.close()
        return redirect(url_for('mostrar_salida'))
    
@mostrar_salida.route('/delete_salida/<string:id>', methods=['POST', 'GET'])
def index_delete_salida(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('DELETE FROM salida_repuesto WHERE id_salida_repuesto = %s', (id,))
    conn.commit()

    cur.close()
    conn.close()
    return redirect(url_for('mostrar_salida'))

def obtener_datos_mysql():
    conn = mysql.connector.connect(**db_config)

    consulta = """
        SELECT r.codigo_repuesto, r.nombre_repuesto, s.cantidad_salida_repuesto, t.codigo_responsable, t.apellidos_responsable, t.nombres_responsable, DATE_FORMAT(s.fecha_salida_repuesto, '%Y-%m-%d')
        FROM salida_repuesto s
        JOIN registrar_repuesto r ON s.id_repuesto = r.id_repuesto
        JOIN registrar_responsable t ON s.id_responsable = t.id_responsable
        """
    datos = pd.read_sql_query(consulta, conn)

    titulos_encabezado = ["CODIGO R", "REPUESTO", "CANTIDAD", "CODIGO R", "APELLIDOS", "NOMBRES", "FECHA"]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Datos'

    for c, titulo in enumerate(titulos_encabezado, start=1):
        cell = worksheet.cell(row=8, column=c)
        cell.value = titulo
        cell.font = Font(bold=True, color="FFFFFFFF", name="Times New Roman")
        cell.fill = PatternFill(fill_type='solid', fgColor='9a5833')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    for r, row in enumerate(datos.values, start=9):
        for c, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r, column=c)
            cell.value = value
            cell.font = Font(name='Times New Roman')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    for column_cells in worksheet.columns:
        column = column_cells[0].column_letter

        if column == 'A':
            worksheet.column_dimensions[column].width = 17
        elif column == 'B':
            worksheet.column_dimensions[column].width = 30
        elif column == 'C':
            worksheet.column_dimensions[column].width = 16
        elif column == 'D':
            worksheet.column_dimensions[column].width = 17
        elif column == 'E':
            worksheet.column_dimensions[column].width = 30
        elif column == 'F':
            worksheet.column_dimensions[column].width = 30
        elif column == 'G':
            worksheet.column_dimensions[column].width = 22


    for r, row_cells in enumerate(worksheet.iter_rows(), start=1):
        adjusted_height = 18
        for cell in row_cells:
            worksheet.row_dimensions[cell.row].height = adjusted_height

    imagen_path = 'src/static/img/ExcelEncabezadoSalidas.png'
    imagen_width = 1130.9
    imagen_height = 166.67

    imagen = Image(imagen_path)
    imagen.width = imagen_width
    imagen.height = imagen_height

    worksheet.add_image(imagen, 'A1')

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        archivo_temporal = tmp.name
        workbook.save(archivo_temporal)

    conn.close()

    return archivo_temporal

@mostrar_salida.route('/generar-excelS', methods=['POST'])
def index_generar_excelS():
    archivo_excel = obtener_datos_mysql()

    return send_file(archivo_excel, as_attachment=True)
    
if __name__ == '__main__':
    mostrar_salida.run(port=3000, debug=True)
