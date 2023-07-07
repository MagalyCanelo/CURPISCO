from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import mysql.connector
import tempfile
import base64

import pandas as pd

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.drawing.image import Image

mostrar_entrada = Flask(__name__)

db_config = {
    'user': 'root',
    'password': '',
    'host': 'localhost',
    'database': 'curpiscobd_repuesto'
}

mostrar_entrada.secret_key = 'mysecretkey'

@mostrar_entrada.route('/mostrar_entrada')
def index_mostrar_entrada():
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute("SELECT e.id_entrada_repuesto, r.codigo_repuesto, r.nombre_repuesto, e.cantidad_entrada_repuesto, p.razon_social_proveedor, e.fecha_entrada_repuesto "
            "FROM entrada_repuesto e "
            "JOIN registrar_repuesto r ON e.id_repuesto = r.id_repuesto "
            "JOIN registrar_proveedor p ON e.id_proveedor = p.id_proveedor")

    results = cur.fetchall()

    cur.close()
    conn.close()
    
    return render_template('MostrarEntrada.html', repuestos=results)

@mostrar_entrada.route('/buscar_entrada', methods=['POST'])
def index_buscar_entrada():
    codigo_repuesto = request.form.get('codigo_repuesto')
    fecha_entrada = request.form.get('fecha_entrada')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify ({'repuesto_entrada': []})

    cur = conn.cursor()

    query = """
    SELECT e.id_entrada_repuesto, r.codigo_repuesto, r.nombre_repuesto, e.cantidad_entrada_repuesto, p.razon_social_proveedor, DATE_FORMAT(e.fecha_entrada_repuesto, '%Y-%m-%d')
    FROM entrada_repuesto e
    JOIN registrar_repuesto r ON e.id_repuesto = r.id_repuesto
    JOIN registrar_proveedor p ON e.id_proveedor = p.id_proveedor
    WHERE 1=1
    """
    params = []

    if codigo_repuesto:
        cur.execute('SELECT id_repuesto FROM registrar_repuesto WHERE codigo_repuesto = %s', (codigo_repuesto,))
        id_repuesto = cur.fetchone()
        if id_repuesto is None:
            conn.close()
            return jsonify({'repuesto_entrada': []})

        id_repuesto = id_repuesto[0]
        query += "AND r.id_repuesto = %s "
        params.append(id_repuesto)

    if fecha_entrada:
        query += "AND DATE(e.fecha_entrada_repuesto) = %s "
        params.append(fecha_entrada)

    cur.execute(query, params)
    data = cur.fetchall()

    conn.close()

    converted_data = []
    for row in data:
        converted_row = []
        for value in row:
            if isinstance(value, bytes):
                converted_row.append(value.decode('utf-8'))
            else:
                converted_row.append(value)
        converted_data.append(converted_row)

    return jsonify({'repuesto_entrada': converted_data})

@mostrar_entrada.route('/limpiar_filtro', methods=['GET'])
def index_limpiar_filtroE():
    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify({'repuesto': []})

    cur = conn.cursor()

    query = """
    SELECT e.id_entrada_repuesto, r.codigo_repuesto, r.nombre_repuesto, e.cantidad_entrada_repuesto, p.razon_social_proveedor, DATE_FORMAT(e.fecha_entrada_repuesto, '%Y-%m-%d')
    FROM entrada_repuesto e
    JOIN registrar_repuesto r ON e.id_repuesto = r.id_repuesto
    JOIN registrar_proveedor p ON e.id_proveedor = p.id_proveedor
    """

    cur.execute(query)
    data = cur.fetchall()

    conn.close()

    converted_data = []
    for row in data:
        converted_row = []
        for value in row:
            if isinstance(value, bytes):
                converted_row.append(value.decode('utf-8'))
            else:
                converted_row.append(value)
        converted_data.append(converted_row)

    return jsonify({'repuesto': converted_data})

#FALTAAA

@mostrar_entrada.route('/edit_entrada/<id>')
def index_edit_entrada(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute("""
        SELECT e.id_entrada_repuesto, r.codigo_repuesto, r.nombre_repuesto, r.descripcion_repuesto, r.foto_repuesto, e.cantidad_entrada_repuesto, p.codigo_proveedor, p.razon_social_proveedor, DATE_FORMAT(e.fecha_entrada_repuesto, '%Y-%m-%d')
        FROM entrada_repuesto e
        JOIN registrar_repuesto r ON e.id_repuesto = r.id_repuesto
        JOIN registrar_proveedor p ON e.id_proveedor = p.id_proveedor
        WHERE e.id_entrada_repuesto = %s
    """, (id,))
    data = cur.fetchall()
    
    column_titles = ['id_repuesto', 'codigo_repuesto', 'nombre_repuesto', 'descripcion_repuesto', 'foto_repuesto', 'cantidad_minima_repuesto', 'codigo_proveedor', 'razon_social_proveedor', 'fecha_entrada_repuesto']
    response_data = {title: [row[i] for row in data] for i, title in enumerate(column_titles)}
    response_data = {key: value[0] if isinstance(value, list) and len(value) == 1 else value for key, value in response_data.items()}
        
    return render_template('editEntrada.html', repuesto=response_data)

@mostrar_entrada.route('/autocompletar_entradaE', methods=['GET', 'POST'])
def autocompletar_repuesto_entradaE():
    codigo_repuesto = request.form.get('codigo_repuesto')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify({'data': {}})
    cursor = conn.cursor()

    query = "SELECT nombre_repuesto, descripcion_repuesto, foto_repuesto FROM registrar_repuesto WHERE codigo_repuesto = %s"
    cursor.execute(query, (codigo_repuesto,))
    result = cursor.fetchone()
    print("Result es:", result)

    if result:
        nombre_repuesto, descripcion_repuesto, foto_repuesto = result

        if foto_repuesto is not None:
            foto_repuesto_encoded = base64.b64encode(foto_repuesto).decode('ascii')
            print("foto_repuesto_encoded:", foto_repuesto_encoded)
        else:
            foto_repuesto_encoded = None

        response_data = {
            'nombre_repuesto': nombre_repuesto,
            'descripcion_repuesto': descripcion_repuesto,
            'foto_repuesto': foto_repuesto_encoded
        }
    else:
        response_data = {}

    cursor.close()
    conn.close()

    return jsonify({'data': response_data})


@mostrar_entrada.route('/autocompletar_proveedorE', methods=['GET', 'POST'])
def autocompletar_proveedor_entradaE():
    codigo_proveedor = request.form.get('codigo_proveedor')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
        return jsonify({'data': {}}) 

    cursor = conn.cursor()
    query = "SELECT razon_social_proveedor FROM registrar_proveedor WHERE codigo_proveedor = %s"
    cursor.execute(query, (codigo_proveedor,))
    result = cursor.fetchone()

    cursor.close()
    conn.close()

    return jsonify({'data': result[0]})

@mostrar_entrada.route('/update_entrada/<id>', methods=['POST'])
def index_update_entrada(id):
    if request.method == "POST":
        codigo_repuesto = request.form['codigo_repuesto']
        cantidad_entrada_repuesto = int(request.form['cantidad_entrada_repuesto'])
        codigo_proveedor = request.form['codigo_proveedor']
        fecha_entrada_repuesto = request.form['fecha_entrada_repuesto']
        
        conn = mysql.connector.connect(**db_config)
        if conn.is_connected():
            print("Conexión exitosa a la base de datos")
        else:
            print("No se pudo establecer la conexión a la base de datos")
        cur = conn.cursor()
        
        cur.execute('SELECT id_repuesto FROM registrar_repuesto WHERE codigo_repuesto = %s', (codigo_repuesto,))
        id_repuesto = cur.fetchone()    
        id_repuesto = id_repuesto[0] 
        cur.execute('SELECT id_proveedor FROM registrar_proveedor WHERE codigo_proveedor = %s', (codigo_proveedor,))
        id_proveedor = cur.fetchone()    
        id_proveedor = id_proveedor[0] 
        
        cur.execute('SELECT cantidad_entrada_repuesto FROM entrada_repuesto WHERE id_entrada_repuesto = %s', (id,))
        cantidad_antes = cur.fetchone()[0] 

        cur.execute('SELECT cantidad_total FROM registrar_repuesto WHERE id_repuesto = %s', (id_repuesto,))
        result = cur.fetchone()
        
        if result:
            cantidad_actual = result[0]
            quitar = cantidad_actual - int(cantidad_antes)
            nueva_cantidad = quitar + cantidad_entrada_repuesto
            cur.execute('UPDATE registrar_repuesto SET cantidad_total = %s WHERE id_repuesto = %s',
            (nueva_cantidad, id_repuesto))
            
            cur.execute("""
                UPDATE entrada_repuesto
                SET id_repuesto =  %s,
                    cantidad_entrada_repuesto = %s,
                    id_proveedor = %s,
                    fecha_entrada_repuesto = %s
                WHERE id_entrada_repuesto  = %s
            """, (id_repuesto, cantidad_entrada_repuesto, id_proveedor, fecha_entrada_repuesto, id))
            conn.commit()

        cur.close()
        conn.close()
        
    return redirect(url_for('mostrar_entrada'))
    
#--------------

@mostrar_entrada.route('/delete_entrada/<string:id>', methods=['POST', 'GET'])
def index_delete_entrada(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('DELETE FROM entrada_repuesto WHERE id_entrada_repuesto = %s', (id,))
    conn.commit()

    cur.close()
    conn.close()
    return redirect(url_for('mostrar_entrada'))

def obtener_datos_mysql():
    conn = mysql.connector.connect(**db_config)

    consulta = """
        SELECT r.codigo_repuesto, r.nombre_repuesto, e.cantidad_entrada_repuesto, p.codigo_proveedor, p.razon_social_proveedor, DATE_FORMAT(e.fecha_entrada_repuesto, '%Y-%m-%d')
        FROM entrada_repuesto e
        JOIN registrar_repuesto r ON e.id_repuesto = r.id_repuesto
        JOIN registrar_proveedor p ON e.id_proveedor = p.id_proveedor
        """
    datos = pd.read_sql_query(consulta, conn)

    titulos_encabezado = ["CODIGO R", "REPUESTO", "CANTIDAD", "CODIGO P", "RAZÓN SOCIAL", "FECHA"]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Datos'

    for c, titulo in enumerate(titulos_encabezado, start=1):
        cell = worksheet.cell(row=8, column=c)
        cell.value = titulo
        cell.font = Font(bold=True, color="FFFFFFFF", name="Times New Roman")
        cell.fill = PatternFill(fill_type='solid', fgColor='9a5833')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    for r, row in enumerate(datos.values, start=9):
        for c, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r, column=c)
            cell.value = value
            cell.font = Font(name='Times New Roman')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    for column_cells in worksheet.columns:
        column = column_cells[0].column_letter

        if column == 'A':
            worksheet.column_dimensions[column].width = 18
        elif column == 'B':
            worksheet.column_dimensions[column].width = 35
        elif column == 'C':
            worksheet.column_dimensions[column].width = 25
        elif column == 'D':
            worksheet.column_dimensions[column].width = 18
        elif column == 'E':
            worksheet.column_dimensions[column].width = 35
        elif column == 'F':
            worksheet.column_dimensions[column].width = 30


    for r, row_cells in enumerate(worksheet.iter_rows(), start=1):
        adjusted_height = 18
        for cell in row_cells:
            worksheet.row_dimensions[cell.row].height = adjusted_height

    imagen_path = 'src/static/img/ExcelEncabezadoEntradas.png'
    imagen_width = 1128.9
    imagen_height = 166.67

    imagen = Image(imagen_path)
    imagen.width = imagen_width
    imagen.height = imagen_height

    worksheet.add_image(imagen, 'A1')

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        archivo_temporal = tmp.name
        workbook.save(archivo_temporal)

    conn.close()

    return archivo_temporal

@mostrar_entrada.route('/generar-excelE', methods=['POST'])
def index_generar_excelE():
    archivo_excel = obtener_datos_mysql()

    return send_file(archivo_excel, as_attachment=True)
    
if __name__ == '__main__':
    mostrar_entrada.run(port=3000, debug=True)
