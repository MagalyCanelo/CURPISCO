from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import mysql.connector
import tempfile

import pandas as pd

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.drawing.image import Image

mostrar_proveedores = Flask(__name__)

db_config = {
    'user': 'root',
    'password':'',
    'host': 'localhost',
    'database': 'curpiscobd_repuesto'
}

mostrar_proveedores.secret_key = 'mysecretkey'

@mostrar_proveedores.route('/')
def index_mostrar_proveedor():
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM registrar_proveedor')
    data = cur.fetchall()
    
    return render_template('MostrarProveedores.html', proveedores=data)

@mostrar_proveedores.route('/limpiar_filtroP', methods=['GET'])
def index_limpiar_filtroP():
    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
    cur = conn.cursor()

    cur.execute('SELECT * FROM registrar_proveedor')
    data = cur.fetchall()

    conn.close()

    converted_data = []
    for value in data:
        converted_row = []
        for val in value:
            if isinstance(val, bytes):
                converted_row.append(val.decode('utf-8'))
            else:
                converted_row.append(val)
        converted_data.append(converted_row)

    return jsonify({'proveedor': converted_data})

@mostrar_proveedores.route('/buscar_proveedor', methods=['POST'])
def index_buscar_proveedor():
    razon_social_proveedor = request.form.get('razon_social_proveedor')
    
    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM registrar_proveedor WHERE razon_social_proveedor LIKE %s', (f'%{razon_social_proveedor}%',))
    data = cur.fetchall()

    conn.close()

    if not data:
        return jsonify({'proveedor': []})

    converted_data = []
    for row in data:
        converted_row = []
        for value in row:
            if isinstance(value, bytes):
                converted_row.append(value.decode('utf-8'))
            else:
                converted_row.append(value)
        converted_data.append(converted_row)
        
    return jsonify({'proveedor': converted_data})

@mostrar_proveedores.route('/edit_proveedor/<id>')
def index_edit_proveedor(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM registrar_proveedor WHERE id_proveedor = %s', (id,))
    data = cur.fetchall()
    
    return render_template('editProveedor.html', proveedor=data[0])

@mostrar_proveedores.route('/update/<id>', methods=['POST'])
def index_update_proveedor(id):
    if request.method == 'POST':
        codigo_proveedor = request.form['codigo_proveedor']
        ruc_proveedor = request.form['ruc_proveedor']
        razon_social_proveedor = request.form['razon_social_proveedor']
        telefono_proveedor = request.form['telefono_proveedor']
        direccion_proveedor = request.form['direccion_proveedor']

        conn = mysql.connector.connect(**db_config)
        cur = conn.cursor()
        cur.execute("""
            UPDATE registrar_proveedor
            SET codigo_proveedor = %s,
                ruc_proveedor = %s,
                razon_social_proveedor = %s,
                telefono_proveedor = %s,
                direccion_proveedor = %s
            WHERE id_proveedor = %s
        """, (codigo_proveedor, ruc_proveedor, razon_social_proveedor, telefono_proveedor, direccion_proveedor, id))
        conn.commit()

        cur.close()
        conn.close()
        return redirect(url_for('mostrar_proveedor'))

@mostrar_proveedores.route('/delete_proveedor/<string:id>', methods=['POST', 'GET'])
def index_delete_proveedor(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('DELETE FROM registrar_proveedor WHERE id_proveedor = %s', (id,))
    conn.commit()

    cur.close()
    conn.close()
    return redirect(url_for('mostrar_proveedor'))

def obtener_datos_mysql():
    conn = mysql.connector.connect(**db_config)

    consulta = "SELECT codigo_proveedor, ruc_proveedor, razon_social_proveedor, telefono_proveedor, direccion_proveedor FROM registrar_proveedor"
    datos = pd.read_sql_query(consulta, conn)

    titulos_encabezado = ["CÓDIGO", "RUC", "RAZÓN SOCIAL", "TELÉFONO", "DIRECCIÓN"]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Datos'

    for c, titulo in enumerate(titulos_encabezado, start=1):
        cell = worksheet.cell(row=8, column=c)
        cell.value = titulo
        cell.font = Font(bold=True, color="FFFFFFFF", name="Times New Roman")
        cell.fill = PatternFill(fill_type='solid', fgColor='9a5833')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    for r, row in enumerate(datos.values, start=9):
        for c, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r, column=c)
            cell.value = value
            cell.font = Font(name='Times New Roman')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    for column_cells in worksheet.columns:
        column = column_cells[0].column_letter

        if column == 'A':
            worksheet.column_dimensions[column].width = 18
        elif column == 'B':
            worksheet.column_dimensions[column].width = 30
        elif column == 'C':
            worksheet.column_dimensions[column].width = 48
        elif column == 'D':
            worksheet.column_dimensions[column].width = 26
        elif column == 'E':
            worksheet.column_dimensions[column].width = 40

    for r, row_cells in enumerate(worksheet.iter_rows(), start=1):
        adjusted_height = 18
        for cell in row_cells:
            worksheet.row_dimensions[cell.row].height = adjusted_height

    imagen_path = 'src/static/img/ExcelEncabezadoProveedores.png'
    imagen_width = 1136.12
    imagen_height = 166.67

    imagen = Image(imagen_path)
    imagen.width = imagen_width
    imagen.height = imagen_height

    worksheet.add_image(imagen, 'A1')

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        archivo_temporal = tmp.name
        workbook.save(archivo_temporal)

    conn.close()

    return archivo_temporal

@mostrar_proveedores.route('/generar-excel', methods=['POST'])
def index_generar_excelP():
    archivo_excel = obtener_datos_mysql()

    return send_file(archivo_excel, as_attachment=True)
    
if __name__ == '__main__':
    mostrar_proveedores.run(port=3000, debug=True)