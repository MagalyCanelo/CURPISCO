from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
import mysql.connector
import tempfile

import pandas as pd

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.drawing.image import Image

mostrar_responsables = Flask(__name__)

db_config = {
    'user': 'root',
    'password': '',
    'host': 'localhost',
    'database': 'curpiscobd_repuesto'
}

mostrar_responsables.secret_key = 'mysecretkey'

@mostrar_responsables.route('/')
def index_mostrar_responsable():
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM registrar_responsable')
    data = cur.fetchall()
    
    return render_template('MostrarResponsables.html', responsables=data)

@mostrar_responsables.route('/limpiar_filtroR', methods=['GET'])
def index_limpiar_filtroR():
    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
    cur = conn.cursor()

    cur.execute('SELECT * FROM registrar_responsable')
    data = cur.fetchall()

    conn.close()

    converted_data = []
    for value in data:
        converted_row = []
        for val in value:
            if isinstance(val, bytes):
                converted_row.append(val.decode('utf-8'))
            else:
                converted_row.append(val)
        converted_data.append(converted_row)

    return jsonify({'responsable': converted_data})

@mostrar_responsables.route('/buscar_responsable', methods=['POST'])
def index_buscar_responsable():
    apellidos_responsable = request.form.get('apellidos_responsable')

    conn = mysql.connector.connect(**db_config)
    if conn.is_connected():
        print("Conexión exitosa a la base de datos")
    else:
        print("No se pudo establecer la conexión a la base de datos")
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM registrar_responsable WHERE apellidos_responsable LIKE %s', (f'%{apellidos_responsable}%',))
    data = cur.fetchall()
    
    conn.close()

    if not data:
        return jsonify({'responsable': []})

    converted_data = []
    for row in data:
        converted_row = []
        for value in row:
            if isinstance(value, bytes):
                converted_row.append(value.decode('utf-8'))
            else:
                converted_row.append(value)
        converted_data.append(converted_row)
        
    return jsonify({'responsable': converted_data})

@mostrar_responsables.route('/edit_responsable/<id>')
def index_edit_responsable(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('SELECT * FROM registrar_responsable WHERE id_responsable = %s', (id,))
    data = cur.fetchall()
    
    return render_template('editResponsable.html', responsable=data[0])

@mostrar_responsables.route('/update/<id>', methods=['POST'])
def index_update_responsable(id):
    if request.method == 'POST':
        codigo_responsable = request.form['codigo_responsable']
        nombres_responsable = request.form['nombres_responsable']
        apellidos_responsable = request.form['apellidos_responsable']
        telefono_responsable = request.form['telefono_responsable']
        direccion_responsable = request.form['direccion_responsable']

        conn = mysql.connector.connect(**db_config)
        cur = conn.cursor()
        
        cur.execute("""
            UPDATE registrar_responsable
            SET codigo_responsable = %s,
                nombres_responsable = %s,
                apellidos_responsable = %s,
                telefono_responsable= %s,
                direccion_responsable = %s
            WHERE id_responsable = %s
        """, (codigo_responsable, nombres_responsable, apellidos_responsable, telefono_responsable, direccion_responsable, id))
        conn.commit()

        cur.close()
        conn.close()
        return redirect(url_for('mostrar_responsable'))

@mostrar_responsables.route('/delete_responsable/<string:id>', methods=['POST', 'GET'])
def index_delete_responsable(id):
    conn = mysql.connector.connect(**db_config)
    cur = conn.cursor()
    
    cur.execute('DELETE FROM registrar_responsable WHERE id_responsable = %s', (id,))
    conn.commit()

    cur.close()
    conn.close()
    return redirect(url_for('mostrar_responsable'))

def obtener_datos_mysql():
    conn = mysql.connector.connect(**db_config)

    consulta = "SELECT codigo_responsable, nombres_responsable, apellidos_responsable, telefono_responsable, direccion_responsable FROM registrar_responsable"
    datos = pd.read_sql_query(consulta, conn)

    titulos_encabezado = ["CÓDIGO", "NOMBRES", "APELLIDOS", "TELÉFONO", "DIRECCIÓN"]

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Datos'

    for c, titulo in enumerate(titulos_encabezado, start=1):
        cell = worksheet.cell(row=8, column=c)
        cell.value = titulo
        cell.font = Font(bold=True, color="FFFFFFFF", name="Times New Roman")
        cell.fill = PatternFill(fill_type='solid', fgColor='9a5833')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    for r, row in enumerate(datos.values, start=9):
        for c, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r, column=c)
            cell.value = value
            cell.font = Font(name='Times New Roman')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

    for column_cells in worksheet.columns:
        column = column_cells[0].column_letter

        if column == 'A':
            worksheet.column_dimensions[column].width = 18
        elif column == 'B':
            worksheet.column_dimensions[column].width = 44
        elif column == 'C':
            worksheet.column_dimensions[column].width = 44
        elif column == 'D':
            worksheet.column_dimensions[column].width = 23
        elif column == 'E':
            worksheet.column_dimensions[column].width = 33

    for r, row_cells in enumerate(worksheet.iter_rows(), start=1):
        adjusted_height = 18
        for cell in row_cells:
            worksheet.row_dimensions[cell.row].height = adjusted_height

    imagen_path = 'src/static/img/ExcelEncabezadoResponsables.png' 
    imagen_width = 1136.12
    imagen_height = 166.67

    imagen = Image(imagen_path)
    imagen.width = imagen_width
    imagen.height = imagen_height

    worksheet.add_image(imagen, 'A1')

    with tempfile.NamedTemporaryFile(delete=False) as tmp:
        archivo_temporal = tmp.name
        workbook.save(archivo_temporal)

    conn.close()

    return archivo_temporal

@mostrar_responsables.route('/generar-excel', methods=['POST'])
def index_generar_excelR():
    archivo_excel = obtener_datos_mysql()

    return send_file(archivo_excel, as_attachment=True)
    
if __name__ == '__main__':
    mostrar_responsables.run(port=3000, debug=True)
